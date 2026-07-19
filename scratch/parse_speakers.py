import json
import subprocess
import sys

# Install openpyxl
subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)

import openpyxl

wb = openpyxl.load_workbook(r"d:\Antigraity Projects\SCD Website final Pranav\Speakers Section\Speakers Data\Speaker.xlsx")

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n=== Sheet: {sheet_name} ===")
    
    # Print all headers (first row)
    headers = []
    for cell in ws[1]:
        headers.append(cell.value)
    print(f"Headers: {headers}")
    print(f"Rows (excluding header): {ws.max_row - 1}")
    print()
    
    # Print all data
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
        print(f"--- Row {row_idx} ---")
        for cell in row:
            col_header = headers[cell.column - 1] if cell.column - 1 < len(headers) else f"Col{cell.column}"
            if cell.value is not None:
                print(f"  {col_header}: {cell.value}")
        print()
